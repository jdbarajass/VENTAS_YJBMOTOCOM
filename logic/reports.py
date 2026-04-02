"""
Generador de reportes para el sistema de rentabilidad.
Maneja la exportación de datos a Excel y otros formatos.
"""

from datetime import date, datetime
from pathlib import Path
from typing import List, Optional
import os

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from database.models import Venta, ResumenDiario, ResumenMensual
from database.db_manager import DatabaseManager
from config import formatear_moneda


class ReportGenerator:
    """
    Generador de reportes en Excel.
    """

    def __init__(self, db_manager: DatabaseManager = None):
        """
        Inicializa el generador de reportes.

        Args:
            db_manager: Instancia del gestor de base de datos
        """
        self.db = db_manager or DatabaseManager()

    def _get_default_path(self, nombre_archivo: str) -> Path:
        """Obtiene la ruta por defecto para guardar reportes."""
        # Usar la carpeta de Documentos del usuario
        documentos = Path.home() / "Documents"
        reportes_dir = documentos / "YJBMOTOCOM_Reportes"
        reportes_dir.mkdir(exist_ok=True)
        return reportes_dir / nombre_archivo

    def exportar_ventas_dia(self, fecha: date, ruta_archivo: Optional[Path] = None) -> Path:
        """
        Exporta las ventas de un día a Excel.

        Args:
            fecha: Fecha de las ventas a exportar
            ruta_archivo: Ruta donde guardar el archivo (opcional)

        Returns:
            Ruta del archivo generado
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("Se requiere pandas y openpyxl para exportar a Excel")

        ventas = self.db.obtener_ventas_por_fecha(fecha)
        config = self.db.obtener_configuracion()
        resumen = self.db.obtener_resumen_diario(fecha, config.gasto_diario)

        if ruta_archivo is None:
            nombre = f"Ventas_{fecha.strftime('%Y-%m-%d')}.xlsx"
            ruta_archivo = self._get_default_path(nombre)

        # Crear DataFrame de ventas
        datos_ventas = []
        for v in ventas:
            datos_ventas.append({
                "Producto": v.producto,
                "Costo": v.costo,
                "Precio Venta": v.precio_venta,
                "Método Pago": v.metodo_pago,
                "Comisión": v.comision,
                "Ganancia Bruta": v.ganancia_bruta,
                "Ganancia Neta": v.ganancia_neta,
                "Notas": v.notas
            })

        df_ventas = pd.DataFrame(datos_ventas)

        # Crear DataFrame de resumen
        datos_resumen = {
            "Métrica": [
                "Total Ventas",
                "Total Costos",
                "Ganancia Bruta",
                "Total Comisiones",
                "Gasto Operativo Diario",
                "UTILIDAD REAL"
            ],
            "Valor": [
                resumen.total_ventas,
                resumen.total_costos,
                resumen.ganancia_bruta,
                resumen.total_comisiones,
                resumen.gasto_operativo,
                resumen.utilidad_real
            ]
        }
        df_resumen = pd.DataFrame(datos_resumen)

        # Escribir a Excel
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
            # Hoja de ventas
            df_ventas.to_excel(writer, sheet_name='Ventas', index=False)

            # Hoja de resumen
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

            # Aplicar formato
            self._aplicar_formato_excel(writer, fecha.strftime("%d/%m/%Y"))

        return ruta_archivo

    def exportar_ventas_mes(self, anio: int, mes: int, ruta_archivo: Optional[Path] = None) -> Path:
        """
        Exporta las ventas de un mes a Excel.

        Args:
            anio: Año
            mes: Mes (1-12)
            ruta_archivo: Ruta donde guardar el archivo (opcional)

        Returns:
            Ruta del archivo generado
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("Se requiere pandas y openpyxl para exportar a Excel")

        ventas = self.db.obtener_ventas_por_mes(anio, mes)
        config = self.db.obtener_configuracion()
        resumen_mes = self.db.obtener_resumen_mensual(anio, mes, config.gasto_diario)
        resumenes_diarios = self.db.obtener_resumen_diario_por_mes(anio, mes, config.gasto_diario)

        nombres_meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        nombre_mes = nombres_meses[mes - 1]

        if ruta_archivo is None:
            nombre = f"Ventas_{nombre_mes}_{anio}.xlsx"
            ruta_archivo = self._get_default_path(nombre)

        # DataFrame de todas las ventas
        datos_ventas = []
        for v in ventas:
            datos_ventas.append({
                "Fecha": v.fecha.strftime("%d/%m/%Y"),
                "Producto": v.producto,
                "Costo": v.costo,
                "Precio Venta": v.precio_venta,
                "Método Pago": v.metodo_pago,
                "Comisión": v.comision,
                "Ganancia Bruta": v.ganancia_bruta,
                "Ganancia Neta": v.ganancia_neta,
                "Notas": v.notas
            })
        df_ventas = pd.DataFrame(datos_ventas)

        # DataFrame de resumen diario
        datos_diarios = []
        for r in resumenes_diarios:
            datos_diarios.append({
                "Fecha": r.fecha.strftime("%d/%m/%Y"),
                "# Ventas": r.num_ventas,
                "Total Ventas": r.total_ventas,
                "Total Costos": r.total_costos,
                "Ganancia Bruta": r.ganancia_bruta,
                "Comisiones": r.total_comisiones,
                "Gasto Operativo": r.gasto_operativo,
                "Utilidad Real": r.utilidad_real,
                "Estado": "Positivo" if r.utilidad_real >= 0 else "Negativo"
            })
        df_diarios = pd.DataFrame(datos_diarios)

        # DataFrame de resumen mensual
        datos_resumen = {
            "Métrica": [
                "Período",
                "Total Ventas",
                "Total Costos",
                "Ganancia Bruta",
                "Total Comisiones",
                "Total Gastos Operativos",
                "UTILIDAD REAL",
                "Número de Ventas",
                "Días con Utilidad Positiva",
                "Días con Utilidad Negativa"
            ],
            "Valor": [
                f"{nombre_mes} {anio}",
                formatear_moneda(resumen_mes.total_ventas),
                formatear_moneda(resumen_mes.total_costos),
                formatear_moneda(resumen_mes.ganancia_bruta),
                formatear_moneda(resumen_mes.total_comisiones),
                formatear_moneda(resumen_mes.total_gastos_operativos),
                formatear_moneda(resumen_mes.utilidad_real),
                str(resumen_mes.num_ventas),
                str(resumen_mes.dias_positivos),
                str(resumen_mes.dias_negativos)
            ]
        }
        df_resumen = pd.DataFrame(datos_resumen)

        # Escribir a Excel
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
            df_ventas.to_excel(writer, sheet_name='Detalle Ventas', index=False)
            df_diarios.to_excel(writer, sheet_name='Resumen Diario', index=False)
            df_resumen.to_excel(writer, sheet_name='Resumen Mensual', index=False)

            self._aplicar_formato_excel(writer, f"{nombre_mes} {anio}")

        return ruta_archivo

    def _aplicar_formato_excel(self, writer, titulo: str):
        """Aplica formato visual al archivo Excel."""
        workbook = writer.book

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar formato a cada hoja
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            # Formato de encabezados
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def abrir_carpeta_reportes(self):
        """Abre la carpeta de reportes en el explorador de archivos."""
        documentos = Path.home() / "Documents"
        reportes_dir = documentos / "YJBMOTOCOM_Reportes"
        reportes_dir.mkdir(exist_ok=True)

        # Abrir en el explorador de archivos
        os.startfile(str(reportes_dir))
