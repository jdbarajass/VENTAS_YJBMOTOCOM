"""
services/exportador.py
Genera archivos .xlsx con los datos de ventas.
Sin dependencias de UI.
"""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

import json
import re as _re

from models.venta import Venta
from models.producto import Producto
from models.factura import Factura
from models.configuracion import Configuracion
from utils.formatters import fecha_corta, nombre_mes


_PAT_TALLA_EXPORT = _re.compile(r"-T:(\w+)$")


def _talla_de(nombre: str) -> str:
    m = _PAT_TALLA_EXPORT.search(nombre or "")
    return m.group(1) if m else "N/A"


# Paleta de colores
_AZUL_HEADER = "1E3A5F"
_AZUL_SUAVE  = "E8F0FE"
_VERDE       = "D4EDDA"
_ROJO        = "F8D7DA"
_GRIS_FILA   = "F5F5F5"


def _borde_fino() -> Border:
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _proteger_hoja(ws) -> None:
    """Marca la hoja como protegida (solo lectura) en Excel. Requiere contraseña para editar."""
    from openpyxl.worksheet.protection import SheetProtection
    ws.protection = SheetProtection(
        sheet=True,
        selectLockedCells=False,
        selectUnlockedCells=False,
        password="YJBMOTOCOM",
    )


# ── Encabezados comunes de ventas ──────────────────────────────────────────
# Col 11 "Pagos JSON" es datos internos — no editar manualmente
_HEADERS_VENTAS = [
    "#", "Fecha", "Producto", "Talla", "Cant.", "Costo", "Precio venta",
    "Método pago", "Comisión", "Ganancia neta", "Notas", "Pagos JSON"
]

_ANCHOS_VENTAS = [5, 12, 30, 8, 7, 15, 16, 14, 14, 15, 28, 1]

_EJEMPLOS_VENTAS = [
    (1, "04/04/2026", "Casco X-Sport T.M",    "M",   1, 85000, 120000, "Efectivo",            0,  35000, "", ""),
    (2, "04/04/2026", "Aceite 10W-40 1L",     "N/A", 2, 18000,  28000, "Transferencia NEQUI", 0,  20000, "", ""),
    (3, "04/04/2026", "Guantes cuero talla L", "N/A", 1, 25000,  40000, "Efectivo",            0, 15000, "Cliente frecuente", ""),
]


def _escribir_encabezados_ventas(ws, titulo_celda: str, titulo_valor: str) -> None:
    """Escribe título (fila 1), fila vacía (2) y encabezados (3) en un worksheet."""
    ws.merge_cells(f"{titulo_celda}:K1")   # solo hasta K — col L es interna
    t = ws[titulo_celda]
    t.value = titulo_valor
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.append([])  # fila 2 vacía

    ws.append(_HEADERS_VENTAS)  # fila 3
    for col_idx in range(1, 13):
        cell = ws.cell(row=3, column=col_idx)
        if col_idx == 12:
            # Columna interna — gris apagada
            cell.font = Font(bold=False, color="AAAAAA", name="Calibri", size=8)
            cell.fill = PatternFill("solid", fgColor="F3F4F6")
        else:
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borde_fino()
    ws.row_dimensions[3].height = 20


def generar_plantilla_ventas_dia(ruta: Path, fecha: date) -> None:
    """
    Genera un .xlsx vacío con el formato correcto para registrar ventas de un día
    y luego importarlo con ⬆ Importar Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas del Día"

    _escribir_encabezados_ventas(
        ws, "A1",
        f"YJBMOTOCOM — Ventas del {fecha_corta(fecha)}"
    )

    # Filas de ejemplo en gris
    for ej in _EJEMPLOS_VENTAS:
        ws.append(list(ej))
        row = ws.max_row
        for col_idx in range(1, 11):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = _borde_fino()
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    # Nota instructiva
    ws.merge_cells(f"A{ws.max_row + 1}:K{ws.max_row + 1}")
    nota = ws.cell(ws.max_row, 1)
    nota.value = (
        "↑ Borra las filas de ejemplo. Agrega tus ventas desde la fila 4. "
        "Si cambias la fecha del título (fila 1), cambia también las fechas de las filas de datos."
    )
    nota.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota.fill = PatternFill("solid", fgColor="FFFBEB")
    nota.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[ws.max_row].height = 24

    for i, ancho in enumerate(_ANCHOS_VENTAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(str(ruta))


def generar_plantilla_ventas_mes(ruta: Path, año: int, mes: int,
                                  prestamos: list | None = None) -> None:
    """
    Genera un .xlsx vacío con el formato correcto para registrar ventas de un mes
    y luego importarlo con ⬆ Importar Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_mes(mes, año)

    _escribir_encabezados_ventas(
        ws, "A1",
        f"YJBMOTOCOM — {nombre_mes(mes, año)}"
    )

    # Filas de ejemplo en gris
    for ej in _EJEMPLOS_VENTAS:
        ws.append(list(ej))
        row = ws.max_row
        for col_idx in range(1, 11):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = _borde_fino()
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    # Nota instructiva
    ws.merge_cells(f"A{ws.max_row + 1}:K{ws.max_row + 1}")
    nota = ws.cell(ws.max_row, 1)
    nota.value = (
        "↑ Borra las filas de ejemplo. Agrega tus ventas desde la fila 4. "
        "Puedes incluir ventas de varios días del mes — cada fila tiene su propia fecha."
    )
    nota.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota.fill = PatternFill("solid", fgColor="FFFBEB")
    nota.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[ws.max_row].height = 24

    for i, ancho in enumerate(_ANCHOS_VENTAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja Préstamos (vacía para que el usuario la llene) ───────────
    ws_prest = wb.create_sheet("Préstamos")
    _escribir_hoja_prestamos(ws_prest, prestamos or [])

    wb.save(str(ruta))


_HEADERS_PRESTAMOS = ["Fecha", "Hora", "Producto", "Almacén", "Observaciones", "Estado"]
_ANCHOS_PRESTAMOS  = [14, 10, 34, 22, 40, 14]
_ESTADOS_COLOR = {
    "pendiente": "FEF3C7",
    "devuelto":  "D1FAE5",
    "cobrado":   "DBEAFE",
}


def _escribir_hoja_prestamos(ws, prestamos: list) -> None:
    """Escribe título, encabezados y datos de préstamos en el worksheet dado."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_PRESTAMOS)

    # Título
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Préstamos"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados
    ws.append(_HEADERS_PRESTAMOS)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    # Filas de datos (o vacío para plantilla)
    for p in prestamos:
        ws.append([
            p.fecha.strftime("%d/%m/%Y") if hasattr(p.fecha, "strftime") else str(p.fecha),
            getattr(p, "hora", "") or "",
            p.producto,
            p.almacen,
            p.observaciones or "",
            p.estado,
        ])
        row = ws.max_row
        color = _ESTADOS_COLOR.get(p.estado, "FFFFFF")
        for col_idx in range(1, ncols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=color)
            c.border = borde
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    # Anchos de columna
    for i, ancho in enumerate(_ANCHOS_PRESTAMOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_INVENTARIO = ["Serial", "Producto", "Talla", "Costo unitario", "Cantidad", "Código barras"]
_ANCHOS_INVENTARIO  = [12, 38, 8, 16, 12, 20]


def _escribir_hoja_inventario(ws, productos: list) -> None:
    """Escribe título, encabezados y datos de inventario en el worksheet dado."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    from datetime import date as _date
    hoy = _date.today().strftime("%d/%m/%Y")

    # Título
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"YJBMOTOCOM — Inventario ({hoy})"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0369A1")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados
    ws.append(_HEADERS_INVENTARIO)
    for col_idx in range(1, 7):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="0284C7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    # Datos
    total_referencias  = 0
    total_unidades     = 0
    total_valor_inv    = 0.0
    for i, p in enumerate(productos, start=1):
        ws.append([
            p.serial or "",
            p.producto,
            p.talla,
            p.costo_unitario,
            p.cantidad,
            p.codigo_barras or "",
        ])
        row = ws.max_row
        fondo = "F0F9FF" if i % 2 == 0 else "FFFFFF"
        for col_idx in range(1, 7):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = borde
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(
                vertical="center",
                horizontal="center" if col_idx in (1, 3, 5, 6) else "left",
            )
        ws.row_dimensions[row].height = 18
        if p.cantidad > 0:
            total_referencias += 1
            total_unidades    += p.cantidad
            total_valor_inv   += p.costo_unitario * p.cantidad

    # ── Fila de totales ───────────────────────────────────────────────────
    ws.append([
        "TOTALES",
        f"{total_referencias} ref. con stock",
        "",
        total_valor_inv,
        total_unidades,
        "",
    ])
    total_row = ws.max_row
    for col_idx in range(1, 7):
        c = ws.cell(row=total_row, column=col_idx)
        c.font = Font(bold=True, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor="E0F2FE")
        c.border = borde
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[total_row].height = 20

    # Anchos
    for i, ancho in enumerate(_ANCHOS_INVENTARIO, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_FACTURAS = ["Descripción", "Proveedor", "Monto", "Fecha llegada", "Fecha vencimiento", "Estado", "Notas", "Fecha pago"]
_ANCHOS_FACTURAS  = [38, 24, 16, 14, 16, 12, 34, 14]
_FACTURA_ESTADO_COLOR = {
    "pendiente": "FEF3C7",
    "pagada":    "DCFCE7",
}


def _escribir_hoja_facturas(ws, facturas: list) -> None:
    """Escribe título, encabezados y datos de facturas en el worksheet dado."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_FACTURAS)

    # Título
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Facturas y Recibos"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="92400E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados
    ws.append(_HEADERS_FACTURAS)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="B45309")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    # Datos
    for i, f in enumerate(facturas, start=1):
        fecha_str = (
            f.fecha_llegada.strftime("%d/%m/%Y")
            if hasattr(f.fecha_llegada, "strftime")
            else str(f.fecha_llegada)
        )
        fv = getattr(f, "fecha_vencimiento", None)
        fv_str = fv.strftime("%d/%m/%Y") if fv else ""
        fp = getattr(f, "fecha_pago", None)
        fp_str = fp.strftime("%d/%m/%Y") if fp else ""
        ws.append([
            f.descripcion,
            f.proveedor,
            f.monto,
            fecha_str,
            fv_str,
            f.estado,
            f.notas or "",
            fp_str,
        ])
        row = ws.max_row
        color = _FACTURA_ESTADO_COLOR.get(f.estado, "FFFFFF")
        for col_idx in range(1, ncols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=color)
            c.border = borde
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    # Anchos
    for i, ancho in enumerate(_ANCHOS_FACTURAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_GASTOS = ["Fecha", "Descripción", "Monto", "Categoría"]
_ANCHOS_GASTOS  = [14, 40, 16, 14]


def _escribir_hoja_gastos(ws, gastos: list) -> None:
    """Escribe título, encabezados y datos de gastos diarios."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Gastos Operativos Diarios"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="6D28D9")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(_HEADERS_GASTOS)
    for col_idx in range(1, 5):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="7C3AED")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    for i, g in enumerate(gastos, start=1):
        fecha_str = (
            g.fecha.strftime("%d/%m/%Y")
            if hasattr(g.fecha, "strftime")
            else str(g.fecha)
        )
        cat = getattr(g, "categoria", "Otro") or "Otro"
        ws.append([fecha_str, g.descripcion, g.monto, cat])
        row = ws.max_row
        fondo = "F5F3FF" if i % 2 == 0 else "FFFFFF"
        for col_idx in range(1, 5):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = borde
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    for i, ancho in enumerate(_ANCHOS_GASTOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_CONFIG = [
    "Arriendo", "Sueldo", "Servicios", "Otros gastos",
    "Días mes", "Comisión Bold (%)", "Comisión Addi (%)", "Comisión Transf. (%)",
    "Modo oscuro", "Inactividad (min)", "Impresora",
]
_ANCHOS_CONFIG = [16, 16, 16, 16, 11, 18, 18, 20, 14, 18, 26]


def _escribir_hoja_configuracion(ws, cfg) -> None:
    """Escribe título, encabezados y fila única de configuración."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_CONFIG)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Configuración"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="065F46")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(_HEADERS_CONFIG)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="047857")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    if cfg is None:
        valores = [0, 0, 0, 0, 30, 0, 0, 0, "No", 10, ""]
    else:
        valores = [
            cfg.arriendo, cfg.sueldo, cfg.servicios, cfg.otros_gastos,
            cfg.dias_mes, cfg.comision_bold, cfg.comision_addi, cfg.comision_transferencia,
            "Sí" if cfg.modo_oscuro else "No", cfg.timeout_minutos, cfg.nombre_impresora or "",
        ]
    ws.append(valores)
    row = ws.max_row
    for col_idx in range(1, ncols + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill = PatternFill("solid", fgColor="ECFDF5")
        c.border = borde
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 22

    for i, ancho in enumerate(_ANCHOS_CONFIG, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_ABONOS = ["Factura", "Proveedor", "Monto abono", "Fecha", "Notas"]
_ANCHOS_ABONOS  = [38, 24, 16, 14, 34]


def _escribir_hoja_abonos(ws, abonos: list) -> None:
    """Escribe título, encabezados y datos de abonos de facturas."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_ABONOS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Abonos de Facturas"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="92400E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(_HEADERS_ABONOS)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="B45309")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    for i, a in enumerate(abonos, start=1):
        fecha_str = (
            a["fecha"].strftime("%d/%m/%Y")
            if hasattr(a["fecha"], "strftime")
            else str(a["fecha"])
        )
        ws.append([
            a["factura_desc"],
            a["factura_prov"],
            a["monto"],
            fecha_str,
            a["notas"],
        ])
        row = ws.max_row
        fondo = "FEF3C7" if i % 2 == 0 else "FFFBEB"
        for col_idx in range(1, ncols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = borde
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    for i, ancho in enumerate(_ANCHOS_ABONOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_NOTAS = ["Tipo", "Texto", "Completado", "Fecha creación", "Fecha límite"]
_ANCHOS_NOTAS  = [16, 52, 12, 18, 14]
_NOTAS_TIPO_LABEL = {"resurtido": "Por Pedir", "tarea": "Tarea"}


def _escribir_hoja_notas(ws, notas: list) -> None:
    """Escribe título, encabezados y datos de notas/pendientes."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_NOTAS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Notas y Pendientes"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="92400E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(_HEADERS_NOTAS)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="B45309")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    for i, n in enumerate(notas, start=1):
        tipo_label = _NOTAS_TIPO_LABEL.get(n.tipo, n.tipo)
        completado_str = "Sí" if n.completado else "No"
        ws.append([tipo_label, n.texto, completado_str, n.fecha_creacion, n.fecha_limite or ""])
        row = ws.max_row
        fondo = "FEF3C7" if not n.completado else "F0FDF4"
        for col_idx in range(1, ncols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = borde
            c.font = Font(name="Calibri", size=10,
                          color="6B7280" if n.completado else "1E293B")
            c.alignment = Alignment(vertical="center",
                                    wrap_text=(col_idx == 2))
        ws.row_dimensions[row].height = 18

    for i, ancho in enumerate(_ANCHOS_NOTAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_USUARIOS = ["Nombre", "Rol"]
_ANCHOS_USUARIOS  = [28, 16]
_ROLES_COLOR = {"admin": "DBEAFE", "vendedor": "F0FDF4"}


def _escribir_hoja_usuarios(ws, usuarios: list) -> None:
    """Escribe título, encabezados y datos de usuarios (nombre + rol, sin contraseñas)."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_USUARIOS)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Usuarios"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1E3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(_HEADERS_USUARIOS)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    for u in usuarios:
        if isinstance(u, dict):
            nombre = str(u.get("nombre", "") or "")
            rol    = str(u.get("rol", "vendedor") or "vendedor")
        else:
            nombre = u.nombre
            rol    = u.rol
        ws.append([nombre, rol])
        row = ws.max_row
        color = _ROLES_COLOR.get(rol, "FFFFFF")
        for col_idx in range(1, ncols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=color)
            c.border = borde
            c.font = Font(name="Calibri", size=10, bold=(rol == "admin"))
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    ws.merge_cells(f"A{ws.max_row + 1}:{get_column_letter(ncols)}{ws.max_row + 1}")
    nota = ws.cell(ws.max_row, 1)
    nota.value = (
        "Roles válidos: admin | vendedor   •   "
        "Las contraseñas NO se exportan por seguridad   •   "
        "Nuevos usuarios importados recibirán clave temporal '1234' — cámbiala en Configuración."
    )
    nota.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota.fill = PatternFill("solid", fgColor="FFFBEB")
    nota.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[ws.max_row].height = 28

    for i, ancho in enumerate(_ANCHOS_USUARIOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


_HEADERS_PRESUPUESTO = ["Año", "Mes", "Categoría", "Monto Presupuestado"]
_ANCHOS_PRESUPUESTO  = [8, 8, 28, 22]


def _escribir_hoja_presupuesto(ws, presupuestos: list) -> None:
    """Escribe título, encabezados y datos de presupuesto mensual por categoría."""
    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    ncols = len(_HEADERS_PRESUPUESTO)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = "YJBMOTOCOM — Presupuesto Mensual"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0E7490")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(_HEADERS_PRESUPUESTO)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="0891B2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[2].height = 20

    for i, p in enumerate(presupuestos, start=1):
        if isinstance(p, dict):
            anio  = p.get("anio", "")
            mes   = p.get("mes", "")
            cat   = p.get("categoria", "")
            monto = p.get("monto_presupuestado", 0)
        else:
            anio  = getattr(p, "anio", "")
            mes   = getattr(p, "mes", "")
            cat   = getattr(p, "categoria", "")
            monto = getattr(p, "monto_presupuestado", 0)
        ws.append([anio, mes, cat, monto])
        row = ws.max_row
        fondo = "ECFEFF" if i % 2 == 0 else "FFFFFF"
        for col_idx in range(1, ncols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = borde
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(
                vertical="center",
                horizontal="right" if col_idx in (1, 2, 4) else "left",
            )
        ws.row_dimensions[row].height = 18

    for i, ancho in enumerate(_ANCHOS_PRESUPUESTO, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def exportar_todo(
    ruta: Path,
    ventas: list | None = None,
    prestamos: list | None = None,
    productos: list | None = None,
    facturas: list | None = None,
    gastos: list | None = None,
    configuracion=None,
    notas: list | None = None,
    abonos: list | None = None,
    usuarios: list | None = None,
    presupuestos: list | None = None,
) -> None:
    """
    Genera un .xlsx con las hojas que se pasen (None = omitir esa hoja).
    Hojas disponibles: Ventas | Préstamos | Inventario | Facturas | Gastos | Configuración
    """
    wb = openpyxl.Workbook()
    primera_hoja_usada = False

    def _hoja(titulo_hoja: str) -> object:
        """Crea o reutiliza la hoja activa (la primera se crea automáticamente)."""
        nonlocal primera_hoja_usada
        if not primera_hoja_usada:
            ws = wb.active
            ws.title = titulo_hoja
            primera_hoja_usada = True
        else:
            ws = wb.create_sheet(titulo_hoja)
        return ws

    # ── Hoja Ventas (opcional) ────────────────────────────────────────────
    if ventas is not None:
        ws_v = _hoja("Ventas")
        ws_v.merge_cells("A1:K1")
        titulo_c = ws_v["A1"]
        titulo_c.value = "YJBMOTOCOM — Historial de Ventas"
        titulo_c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        titulo_c.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
        titulo_c.alignment = Alignment(horizontal="center", vertical="center")
        ws_v.row_dimensions[1].height = 28

        ws_v.append([])
        ws_v.append(_HEADERS_VENTAS)
        for col_idx in range(1, 13):
            cell = ws_v.cell(row=3, column=col_idx)
            if col_idx == 12:
                cell.font = Font(bold=False, color="AAAAAA", name="Calibri", size=8)
                cell.fill = PatternFill("solid", fgColor="F3F4F6")
            else:
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _borde_fino()
        ws_v.row_dimensions[3].height = 20

        total_cant    = 0
        total_costos  = 0.0
        total_ingresos = 0.0
        total_comision = 0.0
        total_neta    = 0.0
        for i, v in enumerate(ventas, start=1):
            pagos_json = json.dumps(v.pagos_combinados, ensure_ascii=False) if v.pagos_combinados else ""
            ws_v.append([
                i, fecha_corta(v.fecha), v.producto, _talla_de(v.producto),
                v.cantidad, v.costo, v.precio, v.metodo_pago,
                v.comision, v.ganancia_neta, v.notas, pagos_json,
            ])
            row = ws_v.max_row
            fondo = _GRIS_FILA if i % 2 == 0 else "FFFFFF"
            for col_idx in range(1, 13):
                c = ws_v.cell(row=row, column=col_idx)
                if col_idx == 12:
                    c.fill = PatternFill("solid", fgColor="F9FAFB")
                    c.font = Font(name="Calibri", size=8, color="AAAAAA")
                else:
                    c.fill = PatternFill("solid", fgColor=fondo)
                    c.font = Font(name="Calibri", size=10)
                c.border = _borde_fino()
            cell_neta = ws_v.cell(row=row, column=10)
            cell_neta.fill = PatternFill(
                "solid", fgColor=_VERDE if v.ganancia_neta >= 0 else _ROJO
            )
            total_cant     += v.cantidad
            total_costos   += v.costo * v.cantidad
            total_ingresos += v.precio * v.cantidad
            total_comision += v.comision
            total_neta     += v.ganancia_neta

        # ── Fila de totales ───────────────────────────────────────────────
        ws_v.append([
            "", "TOTALES",
            f"{len(ventas)} venta(s)",
            "",              # Talla
            total_cant,      # Cant.
            total_costos,    # Costo total
            total_ingresos,  # Ingresos totales
            "",              # Método pago
            total_comision,  # Comisión total
            total_neta,      # Ganancia neta total
            "", "",
        ])
        total_row = ws_v.max_row
        for col_idx in range(1, 13):
            c = ws_v.cell(row=total_row, column=col_idx)
            c.font = Font(bold=True, name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=_AZUL_SUAVE)
            c.border = _borde_fino()
        # Colorear celda de ganancia neta total
        ws_v.cell(row=total_row, column=10).fill = PatternFill(
            "solid", fgColor=_VERDE if total_neta >= 0 else _ROJO
        )

        for i, ancho in enumerate(_ANCHOS_VENTAS, start=1):
            ws_v.column_dimensions[get_column_letter(i)].width = ancho
        ws_v.column_dimensions["L"].width = 1

    # ── Hoja Préstamos (opcional) ─────────────────────────────────────────
    if prestamos is not None:
        _escribir_hoja_prestamos(_hoja("Préstamos"), prestamos)

    # ── Hoja Inventario (opcional) ────────────────────────────────────────
    if productos is not None:
        ws_inv = _hoja("Inventario")
        _escribir_hoja_inventario(ws_inv, productos)
        _proteger_hoja(ws_inv)

    # ── Hoja Facturas (opcional) ──────────────────────────────────────────
    if facturas is not None:
        _escribir_hoja_facturas(_hoja("Facturas"), facturas)

    # ── Hoja Gastos (opcional) ────────────────────────────────────────────
    if gastos is not None:
        _escribir_hoja_gastos(_hoja("Gastos"), gastos)

    # ── Hoja Configuración (opcional) ─────────────────────────────────────
    if configuracion is not None:
        ws_cfg = _hoja("Configuración")
        _escribir_hoja_configuracion(ws_cfg, configuracion)
        _proteger_hoja(ws_cfg)

    # ── Hoja Notas y Pendientes (opcional) ────────────────────────────────
    if notas is not None:
        _escribir_hoja_notas(_hoja("Notas"), notas)

    # ── Hoja Abonos de Facturas (opcional) ────────────────────────────────
    if abonos is not None:
        _escribir_hoja_abonos(_hoja("Abonos"), abonos)

    # ── Hoja Usuarios (opcional, junto con configuración) ─────────────────
    if usuarios is not None:
        _escribir_hoja_usuarios(_hoja("Usuarios"), usuarios)

    # ── Hoja Presupuesto Mensual (opcional) ───────────────────────────────
    if presupuestos is not None:
        _escribir_hoja_presupuesto(_hoja("Presupuesto"), presupuestos)

    # Si ninguna hoja fue incluida, agregar una de aviso
    if not primera_hoja_usada:
        wb.active.title = "Sin datos"
        wb.active["A1"] = "No se seleccionó ninguna hoja para exportar."

    wb.save(str(ruta))


def generar_plantilla_todo(ruta: Path) -> None:
    """
    Genera un .xlsx vacío con las tres hojas (Ventas, Préstamos, Inventario)
    listo para ser rellenado por el usuario e importado desde el panel
    Exportar / Importar.
    Incluye filas de ejemplo en gris para guiar el formato.
    """
    wb = openpyxl.Workbook()

    # ── Hoja Ventas ───────────────────────────────────────────────────────
    ws_v = wb.active
    ws_v.title = "Ventas"
    _escribir_encabezados_ventas(ws_v, "A1", "YJBMOTOCOM — Historial de Ventas")

    for ej in _EJEMPLOS_VENTAS:
        ws_v.append(list(ej))
        row = ws_v.max_row
        for col_idx in range(1, 12):
            c = ws_v.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = _borde_fino()
            c.alignment = Alignment(vertical="center")
        ws_v.row_dimensions[row].height = 18

    ws_v.merge_cells(f"A{ws_v.max_row + 1}:K{ws_v.max_row + 1}")
    nota_v = ws_v.cell(ws_v.max_row, 1)
    nota_v.value = (
        "↑ Borra las filas de ejemplo. Agrega tus ventas desde la fila 4. "
        "Puedes incluir ventas de cualquier mes y año — cada fila tiene su propia fecha."
    )
    nota_v.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_v.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_v.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_v.row_dimensions[ws_v.max_row].height = 24
    for i, ancho in enumerate(_ANCHOS_VENTAS, start=1):
        ws_v.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja Préstamos ────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Préstamos")
    _escribir_hoja_prestamos(ws_p, [])   # vacía con encabezados y formato

    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    _EJEMPLOS_PRESTAMOS = [
        ("15/04/2026", "Casco X-Sport Rojo T.M", "Almacén Norte", "Para revisión", "pendiente"),
        ("20/04/2026", "Guantes cuero talla L",  "Almacén Sur",   "",              "devuelto"),
    ]
    for ej in _EJEMPLOS_PRESTAMOS:
        ws_p.append(list(ej))
        row = ws_p.max_row
        for col_idx in range(1, 6):
            c = ws_p.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde
            c.alignment = Alignment(vertical="center")
        ws_p.row_dimensions[row].height = 18

    ws_p.merge_cells(f"A{ws_p.max_row + 1}:E{ws_p.max_row + 1}")
    nota_p = ws_p.cell(ws_p.max_row, 1)
    nota_p.value = "↑ Borra los ejemplos. Estados válidos: pendiente | devuelto | cobrado"
    nota_p.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_p.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_p.alignment = Alignment(horizontal="center", vertical="center")
    ws_p.row_dimensions[ws_p.max_row].height = 20

    # ── Hoja Inventario ───────────────────────────────────────────────────
    ws_i = wb.create_sheet("Inventario")
    _escribir_hoja_inventario(ws_i, [])   # vacía con encabezados y formato

    _EJEMPLOS_INV = [
        ("001", "Casco X-Sport Rojo T.M",  85000, 5,  "7709001234567"),
        ("002", "Aceite 10W-40 1 litro",   18000, 12, ""),
        ("003", "Guantes cuero talla L",   25000, 8,  "7709009876543"),
    ]
    lado2 = Side(style="thin", color="CCCCCC")
    borde2 = Border(left=lado2, right=lado2, top=lado2, bottom=lado2)
    for ej in _EJEMPLOS_INV:
        ws_i.append(list(ej))
        row = ws_i.max_row
        for col_idx in range(1, 6):
            c = ws_i.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde2
            c.alignment = Alignment(vertical="center")
        ws_i.row_dimensions[row].height = 18

    ws_i.merge_cells(f"A{ws_i.max_row + 1}:E{ws_i.max_row + 1}")
    nota_i = ws_i.cell(ws_i.max_row, 1)
    nota_i.value = "↑ Borra los ejemplos. Agrega tus productos desde la fila 3."
    nota_i.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_i.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_i.alignment = Alignment(horizontal="center", vertical="center")
    ws_i.row_dimensions[ws_i.max_row].height = 20

    # ── Hoja Facturas ─────────────────────────────────────────────────────
    ws_f = wb.create_sheet("Facturas")
    _escribir_hoja_facturas(ws_f, [])   # vacía con encabezados y formato

    _EJEMPLOS_FACT = [
        ("Arriendo local",         "Propietario Norte", 1500000, "01/04/2026", "pendiente", ""),
        ("Proveedor cascos Bogotá", "MotoPartes S.A.S", 850000,  "10/04/2026", "pendiente", "Factura #2341"),
        ("Servicios públicos",      "EPM",               95000,   "05/04/2026", "pagada",    "Pagada el 08/04"),
    ]
    lado3 = Side(style="thin", color="CCCCCC")
    borde3 = Border(left=lado3, right=lado3, top=lado3, bottom=lado3)
    for ej in _EJEMPLOS_FACT:
        ws_f.append(list(ej))
        row = ws_f.max_row
        for col_idx in range(1, 7):
            c = ws_f.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde3
            c.alignment = Alignment(vertical="center")
        ws_f.row_dimensions[row].height = 18

    ws_f.merge_cells(f"A{ws_f.max_row + 1}:F{ws_f.max_row + 1}")
    nota_f = ws_f.cell(ws_f.max_row, 1)
    nota_f.value = "↑ Borra los ejemplos. Estados válidos: pendiente | pagada"
    nota_f.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_f.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_f.alignment = Alignment(horizontal="center", vertical="center")
    ws_f.row_dimensions[ws_f.max_row].height = 20

    # ── Hoja Gastos ───────────────────────────────────────────────────────
    ws_g = wb.create_sheet("Gastos")
    _escribir_hoja_gastos(ws_g, [])

    lado4 = Side(style="thin", color="CCCCCC")
    borde4 = Border(left=lado4, right=lado4, top=lado4, bottom=lado4)
    _EJEMPLOS_GASTOS = [
        ("01/04/2026", "Transporte moto",        25000),
        ("01/04/2026", "Almuerzo",                15000),
        ("02/04/2026", "Insumos limpieza local",  12000),
    ]
    for ej in _EJEMPLOS_GASTOS:
        ws_g.append(list(ej))
        row = ws_g.max_row
        for col_idx in range(1, 4):
            c = ws_g.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde4
            c.alignment = Alignment(vertical="center")
        ws_g.row_dimensions[row].height = 18

    ws_g.merge_cells(f"A{ws_g.max_row + 1}:C{ws_g.max_row + 1}")
    nota_g = ws_g.cell(ws_g.max_row, 1)
    nota_g.value = "↑ Borra los ejemplos. Agrega tus gastos diarios desde la fila 3."
    nota_g.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_g.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_g.alignment = Alignment(horizontal="center", vertical="center")
    ws_g.row_dimensions[ws_g.max_row].height = 20

    # ── Hoja Abonos de Facturas ───────────────────────────────────────────
    ws_a = wb.create_sheet("Abonos")
    _escribir_hoja_abonos(ws_a, [])   # encabezados y título sin datos

    lado6 = Side(style="thin", color="CCCCCC")
    borde6 = Border(left=lado6, right=lado6, top=lado6, bottom=lado6)
    _EJEMPLOS_ABONOS = [
        ("Arriendo local",          "Propietario Norte", 500000,  "01/04/2026", "Primer cuota"),
        ("Proveedor cascos Bogotá", "MotoPartes S.A.S",  250000,  "15/04/2026", ""),
    ]
    for ej in _EJEMPLOS_ABONOS:
        ws_a.append(list(ej))
        row = ws_a.max_row
        for col_idx in range(1, 6):
            c = ws_a.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde6
            c.alignment = Alignment(vertical="center")
        ws_a.row_dimensions[row].height = 18

    ws_a.merge_cells(f"A{ws_a.max_row + 1}:E{ws_a.max_row + 1}")
    nota_a = ws_a.cell(ws_a.max_row, 1)
    nota_a.value = (
        "↑ Borra los ejemplos.  "
        "Factura y Proveedor deben coincidir exactamente con los datos de la hoja Facturas."
    )
    nota_a.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_a.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_a.row_dimensions[ws_a.max_row].height = 24

    # ── Hoja Notas y Pendientes ───────────────────────────────────────────
    ws_n = wb.create_sheet("Notas")
    _escribir_hoja_notas(ws_n, [])   # encabezados y título sin datos

    lado5 = Side(style="thin", color="CCCCCC")
    borde5 = Border(left=lado5, right=lado5, top=lado5, bottom=lado5)
    _EJEMPLOS_NOTAS = [
        ("Por Pedir", "Ej: Cascos XTR-M70 talla M × 5",    "No", "", ""),
        ("Por Pedir", "Ej: Aceite 10W-40 litro × 12",       "Sí", "", ""),
        ("Tarea",     "Ej: Hacer inventario del local",      "No", "", "2026-06-30"),
        ("Tarea",     "Ej: Llamar a proveedor de cascos",    "No", "", ""),
    ]
    for ej in _EJEMPLOS_NOTAS:
        ws_n.append(list(ej))
        row = ws_n.max_row
        for col_idx in range(1, 6):
            c = ws_n.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde5
            c.alignment = Alignment(vertical="center")
        ws_n.row_dimensions[row].height = 18

    ws_n.merge_cells(f"A{ws_n.max_row + 1}:E{ws_n.max_row + 1}")
    nota_n = ws_n.cell(ws_n.max_row, 1)
    nota_n.value = (
        "↑ Borra los ejemplos.  "
        "Tipos válidos: Por Pedir | Tarea   •   "
        "Completado: Sí | No   •   "
        "Fecha límite (opcional): AAAA-MM-DD"
    )
    nota_n.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_n.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_n.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_n.row_dimensions[ws_n.max_row].height = 24

    # ── Hoja Configuración (con valores por defecto del negocio) ─────────
    ws_c = wb.create_sheet("Configuración")
    cfg_defecto = Configuracion(
        arriendo=3_000_000,
        sueldo=2_000_000,
        servicios=300_000,
        otros_gastos=0,
        dias_mes=30,
        comision_bold=0,
        comision_addi=0,
        comision_transferencia=0,
        modo_oscuro=False,
        timeout_minutos=10,
        nombre_impresora="",
    )
    _escribir_hoja_configuracion(ws_c, cfg_defecto)

    ws_c.merge_cells(f"A{ws_c.max_row + 1}:{get_column_letter(len(_HEADERS_CONFIG))}{ws_c.max_row + 1}")
    nota_c = ws_c.cell(ws_c.max_row, 1)
    nota_c.value = (
        "Edita los valores de la fila 3. "
        "Comisiones: porcentaje (ej: 3.49).  "
        "Modo oscuro: Sí | No.  "
        "Inactividad: minutos (ej: 10).  "
        "Impresora: nombre exacto de Windows (puede dejarse vacío)."
    )
    nota_c.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_c.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_c.row_dimensions[ws_c.max_row].height = 28

    # ── Hoja Usuarios ─────────────────────────────────────────────────────
    ws_u = wb.create_sheet("Usuarios")
    _EJEMPLOS_USUARIOS = [
        {"nombre": "Admin",    "rol": "admin"},
        {"nombre": "Vendedor1","rol": "vendedor"},
    ]
    _escribir_hoja_usuarios(ws_u, _EJEMPLOS_USUARIOS)

    # ── Hoja Presupuesto Mensual ───────────────────────────────────────────
    ws_pr = wb.create_sheet("Presupuesto")
    _EJEMPLOS_PRESUPUESTO = [
        {"anio": 2026, "mes": 5, "categoria": "Arriendo",   "monto_presupuestado": 3_000_000},
        {"anio": 2026, "mes": 5, "categoria": "Transporte", "monto_presupuestado": 200_000},
        {"anio": 2026, "mes": 5, "categoria": "Alimentación","monto_presupuestado": 400_000},
    ]
    _escribir_hoja_presupuesto(ws_pr, _EJEMPLOS_PRESUPUESTO)

    lado7 = Side(style="thin", color="CCCCCC")
    borde7 = Border(left=lado7, right=lado7, top=lado7, bottom=lado7)
    for row in ws_pr.iter_rows(min_row=3, max_row=ws_pr.max_row):
        for cell in row:
            if cell.row == ws_pr.max_row:
                break
            cell.fill = PatternFill("solid", fgColor="F1F5F9")
            cell.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            cell.border = borde7

    ws_pr.merge_cells(f"A{ws_pr.max_row + 1}:D{ws_pr.max_row + 1}")
    nota_pr = ws_pr.cell(ws_pr.max_row, 1)
    nota_pr.value = (
        "↑ Borra los ejemplos. Año: AAAA   Mes: 1-12   Categoría: texto libre   Monto: número."
    )
    nota_pr.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota_pr.fill = PatternFill("solid", fgColor="FFFBEB")
    nota_pr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_pr.row_dimensions[ws_pr.max_row].height = 20

    wb.save(str(ruta))


def exportar_ventas_dia(ventas: list[Venta], fecha: date, ruta: Path) -> None:
    """
    Genera un .xlsx con todas las ventas de un día.
    Incluye fila de totales al final.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas del Día"

    # ---- Título ----
    ws.merge_cells("A1:K1")
    titulo = ws["A1"]
    titulo.value = f"YJBMOTOCOM — Ventas del {fecha_corta(fecha)}"
    titulo.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---- Encabezados ----
    headers = [
        "#", "Fecha", "Producto", "Talla", "Cant.", "Costo", "Precio venta",
        "Método pago", "Comisión", "Ganancia neta", "Notas"
    ]
    ws.append([])           # fila 2 vacía
    ws.append(headers)      # fila 3
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borde_fino()
    ws.row_dimensions[3].height = 20

    # ---- Datos ----
    total_ingresos = total_costos = total_comision = total_neta = 0.0
    for i, v in enumerate(ventas, start=1):
        fila = [
            i,
            fecha_corta(v.fecha),
            v.producto,
            _talla_de(v.producto),
            v.cantidad,
            v.costo,
            v.precio,
            v.metodo_pago,
            v.comision,
            v.ganancia_neta,
            v.notas,
        ]
        ws.append(fila)
        row = ws.max_row
        fondo = _GRIS_FILA if i % 2 == 0 else "FFFFFF"
        for col_idx in range(1, 12):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = _borde_fino()
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center")
        # Números a la derecha
        for col_idx in (5, 6, 7, 9, 10):
            ws.cell(row=row, column=col_idx).alignment = Alignment(
                horizontal="right", vertical="center"
            )
        # Color ganancia neta
        cell_neta = ws.cell(row=row, column=10)
        if v.ganancia_neta >= 0:
            cell_neta.fill = PatternFill("solid", fgColor=_VERDE)
        else:
            cell_neta.fill = PatternFill("solid", fgColor=_ROJO)

        total_ingresos += v.precio * v.cantidad
        total_costos += v.costo * v.cantidad
        total_comision += v.comision
        total_neta += v.ganancia_neta

    # ---- Fila totales ----
    ws.append([
        "", "TOTALES", f"{len(ventas)} venta(s)", "",
        "", total_costos, total_ingresos, "",
        total_comision, total_neta, ""
    ])
    total_row = ws.max_row
    for col_idx in range(1, 12):
        c = ws.cell(row=total_row, column=col_idx)
        c.font = Font(bold=True, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=_AZUL_SUAVE)
        c.border = _borde_fino()
    ws.row_dimensions[total_row].height = 18

    # ---- Anchos de columna ----
    anchos = [5, 12, 30, 8, 7, 15, 16, 14, 14, 15, 28]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(str(ruta))


def exportar_ventas_mes(ventas: list[Venta], año: int, mes: int, ruta: Path,
                        prestamos: list | None = None) -> None:
    """
    Genera un .xlsx con todas las ventas de un mes.
    Si se pasan préstamos, agrega una segunda hoja «Préstamos».
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_mes(mes, año)

    ws.merge_cells("A1:K1")
    titulo = ws["A1"]
    titulo.value = f"YJBMOTOCOM — {nombre_mes(mes, año)}"
    titulo.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "#", "Fecha", "Producto", "Talla", "Cant.", "Costo", "Precio venta",
        "Método pago", "Comisión", "Ganancia neta", "Notas"
    ]
    ws.append([])
    ws.append(headers)
    for col_idx in range(1, 12):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borde_fino()

    total_neta = 0.0
    for i, v in enumerate(ventas, start=1):
        ws.append([
            i, fecha_corta(v.fecha), v.producto, _talla_de(v.producto),
            v.cantidad, v.costo, v.precio, v.metodo_pago,
            v.comision, v.ganancia_neta, v.notas,
        ])
        row = ws.max_row
        fondo = _GRIS_FILA if i % 2 == 0 else "FFFFFF"
        for col_idx in range(1, 12):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", fgColor=fondo)
            c.border = _borde_fino()
            c.font = Font(name="Calibri", size=10)
        cell_neta = ws.cell(row=row, column=10)
        if v.ganancia_neta >= 0:
            cell_neta.fill = PatternFill("solid", fgColor=_VERDE)
        else:
            cell_neta.fill = PatternFill("solid", fgColor=_ROJO)
        total_neta += v.ganancia_neta

    ws.append(["", "TOTALES", f"{len(ventas)} venta(s)", "", "", "", "", "", "", total_neta, ""])
    total_row = ws.max_row
    for col_idx in range(1, 12):
        c = ws.cell(row=total_row, column=col_idx)
        c.font = Font(bold=True, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=_AZUL_SUAVE)
        c.border = _borde_fino()

    anchos = [5, 12, 30, 8, 7, 15, 16, 14, 14, 15, 28]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja Préstamos ────────────────────────────────────────────────────
    if prestamos is not None:
        ws_prest = wb.create_sheet("Préstamos")
        _escribir_hoja_prestamos(ws_prest, prestamos)

    wb.save(str(ruta))
