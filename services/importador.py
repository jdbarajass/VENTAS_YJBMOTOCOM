"""
services/importador.py
Importa ventas desde un archivo .xlsx exportado por YJBMOTOCOM.
Sin dependencias de UI.
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
import openpyxl
import re

import json as _json

from models.venta import Venta
from models.prestamo import Prestamo
from models.producto import Producto
from models.factura import Factura
from models.gasto_dia import GastoDia, CATEGORIAS_GASTO
from models.configuracion import Configuracion
from models.cuenta import Cuenta, MovimientoCuenta, CierreMensual
from models.fiado import Fiado, AbonoFiado
from utils.formatters import MESES_ES


# Lookup inverso: "enero" -> 1
_MES_NUM: dict[str, int] = {v.lower(): k for k, v in MESES_ES.items()}


def _leer_prestamos(ws) -> list[Prestamo]:
    """
    Lee la hoja «Préstamos» generada por YJBMOTOCOM.
    Espera: fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Formato nuevo (6 cols): Fecha | Hora | Producto | Almacén | Observaciones | Estado
    Formato viejo (5 cols): Fecha | Producto | Almacén | Observaciones | Estado
    """
    # Detectar formato por el encabezado de la col 2
    header_col2 = str(ws.cell(2, 2).value or "").strip().lower()
    of = 1 if "hora" in header_col2 else 0   # offset para columnas después de Fecha

    prestamos: list[Prestamo] = []
    for row_idx in range(3, ws.max_row + 1):
        producto = str(ws.cell(row_idx, 2 + of).value or "").strip()
        if not producto:
            continue
        almacen = str(ws.cell(row_idx, 3 + of).value or "").strip()
        if not almacen:
            continue

        # Fecha — openpyxl puede devolver datetime (subclase de date)
        fecha_val = ws.cell(row_idx, 1).value
        try:
            if isinstance(fecha_val, datetime):
                p_fecha = fecha_val.date()
            elif isinstance(fecha_val, date):
                p_fecha = fecha_val
            else:
                p_fecha = datetime.strptime(str(fecha_val).strip(), "%d/%m/%Y").date()
        except (ValueError, TypeError):
            p_fecha = date.today()

        hora_val = str(ws.cell(row_idx, 2).value or "").strip() if of == 1 else ""
        observaciones = str(ws.cell(row_idx, 4 + of).value or "").strip()
        estado_raw = str(ws.cell(row_idx, 5 + of).value or "pendiente").strip().lower()
        if estado_raw not in ("pendiente", "devuelto", "cobrado"):
            estado_raw = "pendiente"

        try:
            prestamos.append(Prestamo(
                producto=producto,
                almacen=almacen,
                fecha=p_fecha,
                observaciones=observaciones,
                estado=estado_raw,
                hora=hora_val,
            ))
        except ValueError:
            pass
    return prestamos


@dataclass
class ResultadoImportacionTotal:
    ventas: list[Venta] = field(default_factory=list)
    # Meses detectados desde las fechas de las ventas: {(año, mes), ...}
    meses_afectados: set = field(default_factory=set)
    prestamos: list[Prestamo] = field(default_factory=list)
    productos: list[Producto] = field(default_factory=list)
    facturas: list[Factura] = field(default_factory=list)
    gastos: list[GastoDia] = field(default_factory=list)
    meses_gastos_afectados: set = field(default_factory=set)
    configuracion: Configuracion | None = None
    notas: list | None = None        # None = hoja ausente; [] = hoja vacía; [...] = con datos
    abonos_raw: list | None = None   # None = hoja ausente; [] = hoja vacía; [...] = con datos
    usuarios: list | None = None     # None = hoja ausente; lista de dicts {nombre, rol}
    presupuestos: list | None = None # None = hoja ausente; lista de dicts {anio, mes, categoria, monto_presupuestado}
    cuentas: list[Cuenta] | None = None
    movimientos_cuentas: list | None = None  # lista de dicts {cuenta_nombre, fecha, tipo, monto, descripcion, venta_id}
    cierres_cuentas: list | None = None      # lista de dicts {anio, mes, cuenta_nombre, balance}
    fiados: list[Fiado] | None = None
    abonos_fiado_raw: list | None = None     # lista de dicts {cliente_nombre, descripcion, monto, fecha, notas}
    movimientos_inventario: list | None = None  # lista de dicts {fecha, hora, producto_nombre, tipo, cantidad_ant, cantidad_nva, notas}
    facturas_items_raw: list | None = None      # lista de dicts {factura_desc, factura_prov, descripcion_item, cantidad, precio_unitario}
    log_acciones: list | None = None            # lista de dicts {fecha, hora, usuario, accion, detalle} — se AGREGA, nunca reemplaza
    errores: list[str] = field(default_factory=list)

    @property
    def año(self) -> int | None:
        if not self.meses_afectados:
            return None
        return sorted(self.meses_afectados)[0][0]

    @property
    def mes(self) -> int | None:
        if not self.meses_afectados:
            return None
        return sorted(self.meses_afectados)[0][1]


def _leer_inventario(ws) -> list[Producto]:
    """Lee la hoja «Inventario» generada por exportar_todo."""
    from services.inventario_importador import _detectar_columnas
    productos: list[Producto] = []

    # Detectar fila de encabezados en las primeras 5 filas
    header_row_idx = None
    mapa: dict = {}
    for row_idx in range(1, min(6, ws.max_row + 1)):
        fila = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        celdas = sum(1 for x in fila if x is not None and str(x).strip())
        if celdas < 2:
            continue
        mapa = _detectar_columnas(fila)
        if mapa["producto"] is not None:
            header_row_idx = row_idx
            break

    if header_row_idx is None or mapa["producto"] is None:
        mapa = {"serial": 0, "producto": 1, "costo": 2, "cantidad": 3, "codigo_barras": 4}
        header_row_idx = 1

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        # Saltar filas de totales exportadas por el sistema
        if str(ws.cell(row_idx, 1).value or "").strip().upper() == "TOTALES":
            continue

        col_prod = mapa.get("producto")
        val_prod = ws.cell(row_idx, (col_prod or 0) + 1).value if col_prod is not None else None
        if not val_prod or not str(val_prod).strip():
            continue

        def _get(campo, default):
            idx = mapa.get(campo)
            if idx is None:
                return default
            return ws.cell(row_idx, idx + 1).value

        serial = str(_get("serial", "") or "").strip()
        try:
            costo = float(_get("costo", 0) or 0)
        except (ValueError, TypeError):
            costo = 0.0
        try:
            cantidad = int(float(str(_get("cantidad", 0) or 0).replace(",", ".")))
        except (ValueError, TypeError):
            cantidad = 0
        cod = str(_get("codigo_barras", "") or "").strip()
        cat = str(_get("categoria", "") or "").strip()
        talla = str(_get("talla", "") or "").strip()
        try:
            stock_minimo = int(float(str(_get("stock_minimo", 0) or 0).replace(",", ".")))
        except (ValueError, TypeError):
            stock_minimo = 0

        try:
            productos.append(Producto(
                serial=serial,
                producto=str(val_prod).strip(),
                costo_unitario=costo,
                cantidad=cantidad,
                codigo_barras=cod,
                categoria=cat,
                talla=talla,
                stock_minimo=stock_minimo,
            ))
        except ValueError:
            pass
    return productos


def _leer_facturas(ws) -> list[Factura]:
    """
    Lee la hoja «Facturas» generada por exportar_todo / generar_plantilla_todo.
    Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Descripción | Proveedor | Monto | Fecha llegada | Fecha vencimiento | Estado | Notas
    (Versión anterior sin col 5 también soportada — detecta por encabezado)
    """
    # Detectar si la hoja tiene encabezado de "Fecha vencimiento" en col 5
    header_col5 = str(ws.cell(2, 5).value or "").strip().lower()
    tiene_vencimiento = "venc" in header_col5

    from database.cuentas_repo import obtener_todas_incluyendo_inactivas
    _id_por_nombre = {c.nombre.strip().lower(): c.id for c in obtener_todas_incluyendo_inactivas()}

    facturas: list[Factura] = []
    for row_idx in range(3, ws.max_row + 1):
        descripcion = str(ws.cell(row_idx, 1).value or "").strip()
        if not descripcion:
            continue

        proveedor = str(ws.cell(row_idx, 2).value or "").strip()

        try:
            monto = float(ws.cell(row_idx, 3).value or 0)
        except (ValueError, TypeError):
            monto = 0.0

        fecha_val = ws.cell(row_idx, 4).value
        try:
            if isinstance(fecha_val, datetime):
                fecha_obj = fecha_val.date()
            elif isinstance(fecha_val, date):
                fecha_obj = fecha_val
            else:
                fecha_obj = datetime.strptime(str(fecha_val).strip(), "%d/%m/%Y").date()
        except (ValueError, TypeError):
            fecha_obj = date.today()

        # Columnas desplazadas según versión del archivo
        fecha_pago = None
        if tiene_vencimiento:
            fv_val = ws.cell(row_idx, 5).value
            try:
                if isinstance(fv_val, datetime):
                    fecha_venc = fv_val.date()
                elif isinstance(fv_val, date):
                    fecha_venc = fv_val
                elif fv_val and str(fv_val).strip():
                    fecha_venc = datetime.strptime(str(fv_val).strip(), "%d/%m/%Y").date()
                else:
                    fecha_venc = None
            except (ValueError, TypeError):
                fecha_venc = None
            estado_raw = str(ws.cell(row_idx, 6).value or "pendiente").strip().lower()
            notas = str(ws.cell(row_idx, 7).value or "").strip()
            # Col 8 = fecha_pago (nuevo formato)
            fp_val = ws.cell(row_idx, 8).value
            try:
                if isinstance(fp_val, datetime):
                    fecha_pago = fp_val.date()
                elif isinstance(fp_val, date):
                    fecha_pago = fp_val
                elif fp_val and str(fp_val).strip():
                    fecha_pago = datetime.strptime(str(fp_val).strip(), "%d/%m/%Y").date()
            except (ValueError, TypeError):
                pass
            # Col 9 = Cuenta (nombre) — solo presente en respaldos completos recientes
            cuenta_nombre = str(ws.cell(row_idx, 9).value or "").strip()
        else:
            cuenta_nombre = ""
            fecha_venc = None
            estado_raw = str(ws.cell(row_idx, 5).value or "pendiente").strip().lower()
            notas = str(ws.cell(row_idx, 6).value or "").strip()

        if estado_raw not in ("pendiente", "pagada"):
            estado_raw = "pendiente"

        try:
            facturas.append(Factura(
                descripcion=descripcion,
                proveedor=proveedor,
                monto=monto,
                fecha_llegada=fecha_obj,
                estado=estado_raw,
                notas=notas,
                fecha_vencimiento=fecha_venc,
                fecha_pago=fecha_pago,
                cuenta_id=_id_por_nombre.get(cuenta_nombre.strip().lower()) if cuenta_nombre else None,
            ))
        except ValueError:
            pass
    return facturas


def _leer_gastos(ws) -> list[GastoDia]:
    """
    Lee la hoja «Gastos» generada por exportar_todo.
    Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Fecha | Descripción | Monto | Categoría (col 4, opcional)
    """
    gastos: list[GastoDia] = []
    for row_idx in range(3, ws.max_row + 1):
        fecha_val = ws.cell(row_idx, 1).value
        if fecha_val is None:
            continue
        try:
            if isinstance(fecha_val, datetime):
                fecha_obj = fecha_val.date()
            elif isinstance(fecha_val, date):
                fecha_obj = fecha_val
            else:
                fecha_obj = datetime.strptime(str(fecha_val).strip(), "%d/%m/%Y").date()
        except (ValueError, TypeError):
            continue

        descripcion = str(ws.cell(row_idx, 2).value or "").strip()
        if not descripcion:
            continue
        try:
            monto = float(ws.cell(row_idx, 3).value or 0)
        except (ValueError, TypeError):
            monto = 0.0

        cat_raw = str(ws.cell(row_idx, 4).value or "Otro").strip()
        categoria = cat_raw if cat_raw in CATEGORIAS_GASTO else "Otro"

        try:
            gastos.append(GastoDia(
                descripcion=descripcion,
                monto=monto,
                fecha=fecha_obj,
                categoria=categoria,
            ))
        except ValueError:
            pass
    return gastos


def _leer_configuracion(ws) -> Configuracion | None:
    """
    Lee la hoja «Configuración». Soporta 3 versiones del archivo (detectadas
    por el encabezado de la columna 9):
    - Base (8 cols): solo gastos fijos + comisiones Bold/Addi/Transferencia.
    - Extendida (11 cols): + Modo oscuro | Inactividad (min) | Impresora.
    - Completa (18 cols): + comisiones por cuenta (Nequi/NU/QR/Daviplata/
      Datafono) antes de Modo oscuro, y backup automático al final.
    """
    if ws.max_row < 3:
        return None

    def _num(val, default=0.0):
        try:
            return float(val or default)
        except (ValueError, TypeError):
            return default

    def _int(val, default=30):
        try:
            return int(float(val or default))
        except (ValueError, TypeError):
            return default

    def _bool(val) -> bool:
        return str(val or "no").strip().lower() in ("sí", "si", "yes", "1", "true")

    header_9 = str(ws.cell(2, 9).value or "").strip().lower()
    tiene_comisiones_cuenta = "comisión" in header_9 or "comision" in header_9
    tiene_extendido = bool(header_9)

    try:
        comision_nequi = comision_nu = comision_qr = 0.0
        comision_daviplata = comision_datafono = 0.0
        modo_oscuro = False
        timeout_minutos = 10
        nombre_impresora = ""
        backup_automatico_activo = True
        backup_intervalo_horas = 24

        if tiene_comisiones_cuenta:
            # Formato completo (18 cols)
            comision_nequi     = _num(ws.cell(3, 9).value)
            comision_nu        = _num(ws.cell(3, 10).value)
            comision_qr        = _num(ws.cell(3, 11).value)
            comision_daviplata = _num(ws.cell(3, 12).value)
            comision_datafono  = _num(ws.cell(3, 13).value)
            modo_oscuro = _bool(ws.cell(3, 14).value)
            timeout_minutos = _int(ws.cell(3, 15).value, 10)
            nombre_impresora = str(ws.cell(3, 16).value or "").strip()
            backup_automatico_activo = _bool(ws.cell(3, 17).value)
            backup_intervalo_horas = _int(ws.cell(3, 18).value, 24)
        elif tiene_extendido:
            # Formato extendido (11 cols, sin comisiones por cuenta ni backup)
            modo_oscuro = _bool(ws.cell(3, 9).value)
            timeout_minutos = _int(ws.cell(3, 10).value, 10)
            nombre_impresora = str(ws.cell(3, 11).value or "").strip()

        return Configuracion(
            arriendo=_num(ws.cell(3, 1).value),
            sueldo=_num(ws.cell(3, 2).value),
            servicios=_num(ws.cell(3, 3).value),
            otros_gastos=_num(ws.cell(3, 4).value),
            dias_mes=_int(ws.cell(3, 5).value),
            comision_bold=_num(ws.cell(3, 6).value),
            comision_addi=_num(ws.cell(3, 7).value),
            comision_transferencia=_num(ws.cell(3, 8).value),
            comision_nequi=comision_nequi,
            comision_nu=comision_nu,
            comision_qr=comision_qr,
            comision_daviplata=comision_daviplata,
            comision_datafono=comision_datafono,
            modo_oscuro=modo_oscuro,
            timeout_minutos=timeout_minutos,
            nombre_impresora=nombre_impresora,
            backup_automatico_activo=backup_automatico_activo,
            backup_intervalo_horas=backup_intervalo_horas,
        )
    except Exception:
        return None


def _leer_usuarios(ws) -> list[dict]:
    """
    Lee la hoja «Usuarios».
    Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Nombre | Rol
    Ignora la fila de nota instructiva que el exportador escribe al final
    (texto largo, sin valor real de "Rol" propio de un usuario).
    Retorna lista de dicts {nombre, rol}.
    """
    usuarios: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        nombre = str(ws.cell(row_idx, 1).value or "").strip()
        if not nombre:
            continue
        if len(nombre) > 50 or nombre.lower().startswith("roles válidos"):
            continue
        rol_raw = str(ws.cell(row_idx, 2).value or "vendedor").strip().lower()
        rol = "admin" if rol_raw == "admin" else "vendedor"
        usuarios.append({"nombre": nombre, "rol": rol})
    return usuarios


def _leer_presupuesto(ws) -> list[dict]:
    """
    Lee la hoja «Presupuesto» generada por exportar_todo() / generar_plantilla_todo().
    Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Año | Mes | Categoría | Monto Presupuestado
    """
    presupuestos: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        anio_val = ws.cell(row_idx, 1).value
        if anio_val is None:
            continue
        try:
            anio = int(anio_val)
        except (ValueError, TypeError):
            continue
        mes_val = ws.cell(row_idx, 2).value
        try:
            mes = int(mes_val)
        except (ValueError, TypeError):
            continue
        if not (1 <= mes <= 12):
            continue
        categoria = str(ws.cell(row_idx, 3).value or "").strip()
        if not categoria:
            continue
        try:
            monto = float(ws.cell(row_idx, 4).value or 0)
        except (ValueError, TypeError):
            monto = 0.0
        presupuestos.append({
            "anio": anio,
            "mes": mes,
            "categoria": categoria,
            "monto_presupuestado": monto,
        })
    return presupuestos


def _leer_abonos(ws) -> list[dict]:
    """
    Lee la hoja «Abonos» generada por exportar_todo().
    Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Factura | Proveedor | Monto abono | Fecha | Notas
    Retorna lista de dicts con claves: factura_desc, factura_prov, monto, fecha, notas.
    El factura_id se resuelve durante _ejecutar_importacion() por matching.
    """
    abonos: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        factura_desc = str(ws.cell(row_idx, 1).value or "").strip()
        if not factura_desc:
            continue
        factura_prov = str(ws.cell(row_idx, 2).value or "").strip()
        try:
            monto = float(ws.cell(row_idx, 3).value or 0)
        except (ValueError, TypeError):
            monto = 0.0
        if monto <= 0:
            continue
        fecha_val = ws.cell(row_idx, 4).value
        try:
            if isinstance(fecha_val, datetime):
                fecha_obj = fecha_val.date()
            elif isinstance(fecha_val, date):
                fecha_obj = fecha_val
            else:
                fecha_obj = datetime.strptime(str(fecha_val).strip(), "%d/%m/%Y").date()
        except (ValueError, TypeError):
            fecha_obj = date.today()
        notas = str(ws.cell(row_idx, 5).value or "").strip()
        abonos.append({
            "factura_desc": factura_desc,
            "factura_prov": factura_prov,
            "monto": monto,
            "fecha": fecha_obj,
            "notas": notas,
        })
    return abonos


_RE_HORA = re.compile(r"^\d{1,2}:\d{2}(:\d{2})?$")

# Mapeo inverso: etiqueta legible del Excel → tipo interno en BD
_NOTAS_LABEL_A_TIPO = {
    "por pedir": "resurtido",
    "resurtido": "resurtido",
    "tarea":     "tarea",
}


def _leer_notas(ws):
    """
    Lee la hoja «Notas» generada por exportar_todo().
    Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Tipo | Texto | Completado | Fecha creación | Fecha límite (opcional)
    """
    from datetime import datetime
    from models.nota import Nota
    notas: list[Nota] = []
    for row_idx in range(3, ws.max_row + 1):
        tipo_raw = str(ws.cell(row_idx, 1).value or "").strip().lower()
        tipo = _NOTAS_LABEL_A_TIPO.get(tipo_raw)
        if tipo is None:
            continue
        texto = str(ws.cell(row_idx, 2).value or "").strip()
        if not texto:
            continue
        completado_raw = str(ws.cell(row_idx, 3).value or "No").strip().lower()
        completado = completado_raw in ("sí", "si", "yes", "1", "true")
        fecha_creacion = str(ws.cell(row_idx, 4).value or "").strip()
        if not fecha_creacion:
            fecha_creacion = datetime.now().strftime("%Y-%m-%d %H:%M")
        fecha_limite_raw = str(ws.cell(row_idx, 5).value or "").strip()
        fecha_limite = fecha_limite_raw if fecha_limite_raw else None
        notas.append(Nota(
            texto=texto,
            tipo=tipo,
            completado=completado,
            fecha_creacion=fecha_creacion,
            fecha_limite=fecha_limite,
        ))
    return notas


def _leer_cuentas(ws) -> list[Cuenta]:
    """
    Lee la hoja «Cuentas». Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Nombre | Método Pago | Balance Actual | Color | Activa | Orden
    """
    cuentas: list[Cuenta] = []
    for row_idx in range(3, ws.max_row + 1):
        nombre = str(ws.cell(row_idx, 1).value or "").strip()
        if not nombre:
            continue
        metodo_pago = str(ws.cell(row_idx, 2).value or "").strip()
        try:
            balance = float(ws.cell(row_idx, 3).value or 0)
        except (ValueError, TypeError):
            balance = 0.0
        color = str(ws.cell(row_idx, 4).value or "#3B82F6").strip()
        activa_raw = str(ws.cell(row_idx, 5).value or "Sí").strip().lower()
        activa = activa_raw in ("sí", "si", "yes", "1", "true")
        try:
            orden = int(ws.cell(row_idx, 6).value or 0)
        except (ValueError, TypeError):
            orden = 0
        cuentas.append(Cuenta(
            nombre=nombre, metodo_pago=metodo_pago, balance_actual=balance,
            color=color, activa=activa, orden=orden,
        ))
    return cuentas


def _leer_movimientos_cuentas(ws) -> list[dict]:
    """
    Lee la hoja «Mov. Cuentas». Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: ID | Cuenta (nombre) | Fecha | Tipo | Monto | Descripción | Venta ID
    """
    _TIPO_LABEL_A_RAW = {
        "venta": "venta", "ajuste manual": "ajuste_manual",
        "transferencia salida": "transferencia_salida",
        "transferencia entrada": "transferencia_entrada",
        "gasto operativo": "gasto_operativo", "reversión gasto": "reversa_gasto",
    }
    movimientos: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        cuenta_nombre = str(ws.cell(row_idx, 2).value or "").strip()
        if not cuenta_nombre:
            continue
        fecha = str(ws.cell(row_idx, 3).value or "").strip()
        tipo_raw = str(ws.cell(row_idx, 4).value or "").strip().lower()
        tipo = _TIPO_LABEL_A_RAW.get(tipo_raw, tipo_raw)
        try:
            monto = float(ws.cell(row_idx, 5).value or 0)
        except (ValueError, TypeError):
            monto = 0.0
        descripcion = str(ws.cell(row_idx, 6).value or "").strip()
        venta_id_raw = ws.cell(row_idx, 7).value
        try:
            venta_id = int(venta_id_raw) if venta_id_raw else None
        except (ValueError, TypeError):
            venta_id = None
        movimientos.append({
            "cuenta_nombre": cuenta_nombre, "fecha": fecha, "tipo": tipo,
            "monto": monto, "descripcion": descripcion, "venta_id": venta_id,
        })
    return movimientos


def _leer_cierres_cuentas(ws) -> list[dict]:
    """
    Lee la hoja «Cierres Cuentas». Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Año | Mes | Cuenta (nombre) | Balance al Cierre | Fecha Cierre | Notas
    Cada fila es UN registro de cuenta dentro de un cierre — varias filas con el
    mismo Año/Mes pertenecen al mismo cierre mensual (se reagrupan al guardar).
    """
    cierres: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        anio_raw = ws.cell(row_idx, 1).value
        if not anio_raw:
            continue
        try:
            anio = int(anio_raw)
            mes = int(ws.cell(row_idx, 2).value or 0)
        except (ValueError, TypeError):
            continue
        cuenta_nombre = str(ws.cell(row_idx, 3).value or "").strip()
        try:
            balance = float(ws.cell(row_idx, 4).value or 0)
        except (ValueError, TypeError):
            balance = 0.0
        fecha_cierre = str(ws.cell(row_idx, 5).value or "").strip()
        notas = str(ws.cell(row_idx, 6).value or "").strip()
        cierres.append({
            "anio": anio, "mes": mes, "cuenta_nombre": cuenta_nombre,
            "balance": balance, "fecha_cierre": fecha_cierre, "notas": notas,
        })
    return cierres


def _leer_fiado(ws) -> list[Fiado]:
    """
    Lee la hoja «Fiado». Fila 1 = título, fila 2 = vacía, fila 3 = encabezados, fila 4+ = datos.
    Columnas: # | Cliente | Cédula | Teléfono | Descripción | Deuda total | Fecha | Estado | Notas
    """
    fiados: list[Fiado] = []
    for row_idx in range(4, ws.max_row + 1):
        cliente = str(ws.cell(row_idx, 2).value or "").strip()
        if not cliente:
            continue
        cedula = str(ws.cell(row_idx, 3).value or "").strip()
        tel = str(ws.cell(row_idx, 4).value or "").strip()
        descripcion = str(ws.cell(row_idx, 5).value or "").strip()
        try:
            monto_total = float(ws.cell(row_idx, 6).value or 0)
        except (ValueError, TypeError):
            monto_total = 0.0
        fecha_val = ws.cell(row_idx, 7).value
        try:
            if isinstance(fecha_val, datetime):
                fecha_obj = fecha_val.date()
            elif isinstance(fecha_val, date):
                fecha_obj = fecha_val
            else:
                fecha_obj = datetime.strptime(str(fecha_val).strip(), "%d/%m/%Y").date()
        except (ValueError, TypeError):
            fecha_obj = date.today()
        estado = str(ws.cell(row_idx, 8).value or "pendiente").strip().lower()
        if estado not in ("pendiente", "pagado"):
            estado = "pendiente"
        notas = str(ws.cell(row_idx, 9).value or "").strip()
        try:
            fiados.append(Fiado(
                cliente_nombre=cliente, cliente_cedula=cedula, cliente_tel=tel,
                descripcion=descripcion, monto_total=monto_total, fecha=fecha_obj,
                estado=estado, notas=notas,
            ))
        except ValueError:
            pass
    return fiados


def _leer_abonos_fiado(ws) -> list[dict]:
    """
    Lee la hoja «Abonos Fiado». Fila 1 = título, fila 2 = vacía, fila 3 = encabezados, fila 4+ = datos.
    Columnas: # | Cliente | Descripción deuda | Monto abono | Fecha | Notas
    """
    abonos: list[dict] = []
    for row_idx in range(4, ws.max_row + 1):
        cliente = str(ws.cell(row_idx, 2).value or "").strip()
        if not cliente:
            continue
        descripcion = str(ws.cell(row_idx, 3).value or "").strip()
        try:
            monto = float(ws.cell(row_idx, 4).value or 0)
        except (ValueError, TypeError):
            monto = 0.0
        fecha_val = ws.cell(row_idx, 5).value
        try:
            if isinstance(fecha_val, datetime):
                fecha_obj = fecha_val.date()
            elif isinstance(fecha_val, date):
                fecha_obj = fecha_val
            else:
                fecha_obj = datetime.strptime(str(fecha_val).strip(), "%d/%m/%Y").date()
        except (ValueError, TypeError):
            fecha_obj = date.today()
        notas = str(ws.cell(row_idx, 6).value or "").strip()
        abonos.append({
            "cliente_nombre": cliente, "descripcion": descripcion,
            "monto": monto, "fecha": fecha_obj, "notas": notas,
        })
    return abonos


def _leer_movimientos_inventario(ws) -> list[dict]:
    """
    Lee la hoja «Mov. Inventario». Fila 1 = título, fila 2 = vacía, fila 3 = encabezados, fila 4+ = datos.
    Columnas: # | Fecha | Hora | Producto | Tipo | Cantidad anterior | Cantidad nueva | Cambio | Notas
    """
    movimientos: list[dict] = []
    for row_idx in range(4, ws.max_row + 1):
        producto = str(ws.cell(row_idx, 4).value or "").strip()
        if not producto:
            continue
        fecha = str(ws.cell(row_idx, 2).value or "").strip()
        hora = str(ws.cell(row_idx, 3).value or "").strip()
        tipo = str(ws.cell(row_idx, 5).value or "").strip()
        try:
            cant_ant = int(float(ws.cell(row_idx, 6).value or 0))
        except (ValueError, TypeError):
            cant_ant = 0
        try:
            cant_nva = int(float(ws.cell(row_idx, 7).value or 0))
        except (ValueError, TypeError):
            cant_nva = 0
        notas = str(ws.cell(row_idx, 9).value or "").strip()
        movimientos.append({
            "fecha": fecha, "hora": hora, "producto_nombre": producto, "tipo": tipo,
            "cantidad_ant": cant_ant, "cantidad_nva": cant_nva, "notas": notas,
        })
    return movimientos


def _leer_facturas_items(ws) -> list[dict]:
    """
    Lee la hoja «Facturas Items». Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Factura | Proveedor | Descripción Item | Cantidad | Precio Unitario | Subtotal
    """
    items: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        factura_desc = str(ws.cell(row_idx, 1).value or "").strip()
        if not factura_desc:
            continue
        factura_prov = str(ws.cell(row_idx, 2).value or "").strip()
        descripcion_item = str(ws.cell(row_idx, 3).value or "").strip()
        try:
            cantidad = float(ws.cell(row_idx, 4).value or 0)
        except (ValueError, TypeError):
            cantidad = 0.0
        try:
            precio_unitario = float(ws.cell(row_idx, 5).value or 0)
        except (ValueError, TypeError):
            precio_unitario = 0.0
        items.append({
            "factura_desc": factura_desc, "factura_prov": factura_prov,
            "descripcion_item": descripcion_item,
            "cantidad": cantidad, "precio_unitario": precio_unitario,
        })
    return items


def _leer_log_acciones(ws) -> list[dict]:
    """
    Lee la hoja «Log Auditoría». Fila 1 = título, fila 2 = encabezados, fila 3+ = datos.
    Columnas: Fecha | Hora | Usuario | Acción | Detalle
    """
    registros: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        fecha = str(ws.cell(row_idx, 1).value or "").strip()
        if not fecha:
            continue
        registros.append({
            "fecha": fecha,
            "hora": str(ws.cell(row_idx, 2).value or "").strip(),
            "usuario": str(ws.cell(row_idx, 3).value or "").strip(),
            "accion": str(ws.cell(row_idx, 4).value or "").strip(),
            "detalle": str(ws.cell(row_idx, 5).value or "").strip(),
        })
    return registros


_RE_SOLO_NUM = re.compile(r"^\d+(\.\d+)?$")
_METODOS_BASE = {"Efectivo", "Addi", "Datafono", "Transferencia", "Combinado", "Otro", "Bold"}


def validar_resultado(res: "ResultadoImportacionTotal") -> tuple[list[str], list[str]]:
    """
    Verifica coherencia de los datos leídos del Excel antes de escribir en BD.
    Retorna (errores_criticos, advertencias).
    Errores críticos bloquean la importación; advertencias solo informan.
    """
    errores: list[str] = []
    adv: list[str] = []

    # ── Préstamos ──────────────────────────────────────────────────────────
    # Detectar columnas transpuestas: producto o almacén con formato HH:MM
    sospechosos_p = [
        p for p in res.prestamos
        if _RE_HORA.match(p.producto) or _RE_HORA.match(p.almacen)
    ]
    if sospechosos_p:
        ejemplo = sospechosos_p[0].producto
        errores.append(
            f"Préstamos ({len(sospechosos_p)} fila(s)): el campo 'Producto' contiene "
            f"'{ejemplo}' — parece una hora, columnas transpuestas. "
            "Exporta de nuevo con la versión actualizada del sistema."
        )

    # ── Ventas ─────────────────────────────────────────────────────────────
    if res.ventas:
        n = len(res.ventas)

        # Más del 50% con precio = 0 → probable desplazamiento de columnas
        precio_cero = sum(1 for v in res.ventas if v.precio == 0)
        if precio_cero > n * 0.5:
            errores.append(
                f"Ventas ({precio_cero}/{n} registros): el campo 'Precio' es 0 en la mayoría "
                "de filas — posible desplazamiento de columnas."
            )

        # Más del 50% con costo = 0 (menos grave, puede ser deliberado, solo advierte)
        costo_cero = sum(1 for v in res.ventas if v.costo == 0)
        if costo_cero > n * 0.5:
            adv.append(
                f"Ventas ({costo_cero}/{n} registros): el campo 'Costo' es 0 en la mayoría "
                "de filas — verifica que los costos estén correctos."
            )

        # Métodos de pago desconocidos (primera palabra)
        metodos_inv = {
            v.metodo_pago for v in res.ventas
            if v.metodo_pago.split()[0] not in _METODOS_BASE
        }
        if metodos_inv:
            adv.append(
                f"Ventas: métodos de pago no reconocidos → "
                + ", ".join(sorted(metodos_inv)[:4])
            )

        # Sin meses detectados cuando hay filas → fechas mal formateadas
        if not res.meses_afectados:
            adv.append("Ventas: no se detectaron meses válidos; las fechas podrían estar mal formateadas.")

    # ── Inventario ─────────────────────────────────────────────────────────
    if res.productos:
        n = len(res.productos)

        # Nombres de producto puramente numéricos → probable columna desplazada
        prod_num = [p for p in res.productos if _RE_SOLO_NUM.match(p.producto)]
        if len(prod_num) > n * 0.3:
            errores.append(
                f"Inventario ({len(prod_num)}/{n} productos): el campo 'Producto' "
                f"contiene valores numéricos (ej: '{prod_num[0].producto}') — "
                "posible desplazamiento de columnas."
            )

        # Todos los costos en cero
        if all(p.costo_unitario == 0 for p in res.productos):
            adv.append("Inventario: todos los productos tienen costo = 0.")

    # ── Facturas ───────────────────────────────────────────────────────────
    if res.facturas:
        n = len(res.facturas)
        monto_cero = sum(1 for f in res.facturas if f.monto == 0)
        if monto_cero > n * 0.5:
            adv.append(
                f"Facturas ({monto_cero}/{n}): la mayoría tienen monto = 0 — "
                "verifica que los montos estén en la columna correcta."
            )

    # ── Gastos ─────────────────────────────────────────────────────────────
    negativos = [g for g in res.gastos if g.monto < 0]
    if negativos:
        adv.append(f"Gastos: {len(negativos)} registro(s) con monto negativo.")

    return errores, adv


def importar_todo(ruta: Path) -> ResultadoImportacionTotal:
    """
    Lee un .xlsx exportado por exportar_todo() y devuelve ventas, préstamos
    e inventario parseados.
    Espera hojas: «Ventas», «Préstamos», «Inventario».
    """
    resultado = ResultadoImportacionTotal()

    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
    except Exception as exc:
        resultado.errores.append(f"No se pudo abrir el archivo: {exc}")
        return resultado

    # ── Hoja Ventas ────────────────────────────────────────────────────────
    ws_v = next(
        (wb[s] for s in wb.sheetnames if s.lower() == "ventas"),
        None,
    )
    if ws_v is None:
        resultado.errores.append("No se encontró la hoja 'Ventas' en el archivo.")
    else:
        # Detectar formato nuevo (col 4 = "Talla") vs. formato viejo
        _hdr4 = str(ws_v.cell(3, 4).value or "").strip().lower()
        _of = 1 if "talla" in _hdr4 else 0  # offset para columnas después de Producto

        for row_idx in range(4, ws_v.max_row + 1):
            col2 = ws_v.cell(row_idx, 2).value
            if col2 is None:
                continue
            if str(col2).upper().strip() == "TOTALES":
                break
            try:
                if isinstance(col2, datetime):
                    venta_fecha = col2.date()
                elif isinstance(col2, date):
                    venta_fecha = col2
                else:
                    venta_fecha = datetime.strptime(str(col2).strip(), "%d/%m/%Y").date()
            except (ValueError, TypeError):
                resultado.errores.append(f"Fila {row_idx}: fecha inválida — omitida")
                continue

            producto = str(ws_v.cell(row_idx, 3).value or "").strip()
            if not producto:
                continue
            try:
                cantidad = int(ws_v.cell(row_idx, 4 + _of).value or 1)
                if cantidad < 1:
                    cantidad = 1
            except (ValueError, TypeError):
                cantidad = 1
            try:
                costo = float(ws_v.cell(row_idx, 5 + _of).value or 0)
            except (ValueError, TypeError):
                costo = 0.0
            try:
                precio = float(ws_v.cell(row_idx, 6 + _of).value or 0)
            except (ValueError, TypeError):
                precio = 0.0
            metodo = str(ws_v.cell(row_idx, 7 + _of).value or "Efectivo").strip() or "Efectivo"
            try:
                comision = float(ws_v.cell(row_idx, 8 + _of).value or 0)
            except (ValueError, TypeError):
                comision = 0.0
            try:
                ganancia = float(ws_v.cell(row_idx, 9 + _of).value or 0)
            except (ValueError, TypeError):
                ganancia = 0.0
            notas = str(ws_v.cell(row_idx, 10 + _of).value or "").strip()
            # Col 11/12: pagos_combinados JSON (solo existe en archivos exportados)
            pagos_raw = ws_v.cell(row_idx, 11 + _of).value
            pagos_combinados = None
            if pagos_raw:
                try:
                    pagos_combinados = _json.loads(str(pagos_raw))
                except Exception:
                    pass

            # Columnas extendidas (solo presentes en respaldos completos generados
            # con exportar_todo desde esta versión en adelante; archivos viejos o
            # plantillas manuales simplemente no las tienen — quedan en sus valores
            # por defecto).
            _base = 11 + _of  # misma base que pagos_raw (col "Pagos JSON")
            hora = str(ws_v.cell(row_idx, _base + 1).value or "").strip()
            vendedor = str(ws_v.cell(row_idx, _base + 2).value or "").strip()
            cliente_nombre = str(ws_v.cell(row_idx, _base + 3).value or "").strip()
            cliente_cedula = str(ws_v.cell(row_idx, _base + 4).value or "").strip()
            cliente_tel = str(ws_v.cell(row_idx, _base + 5).value or "").strip()
            try:
                numero_factura = int(ws_v.cell(row_idx, _base + 6).value or 0) or None
            except (ValueError, TypeError):
                numero_factura = None
            try:
                descuento = int(ws_v.cell(row_idx, _base + 7).value or 0)
            except (ValueError, TypeError):
                descuento = 0
            sku = str(ws_v.cell(row_idx, _base + 8).value or "").strip()
            try:
                precio_ofertado = float(ws_v.cell(row_idx, _base + 9).value or 0)
            except (ValueError, TypeError):
                precio_ofertado = 0.0
            try:
                grupo_venta_id = int(ws_v.cell(row_idx, _base + 10).value or 0) or None
            except (ValueError, TypeError):
                grupo_venta_id = None
            # Leer talla de col 4 solo cuando el archivo tiene ese formato (_of == 1)
            talla_excel = str(ws_v.cell(row_idx, 4).value or "").strip() if _of == 1 else ""

            try:
                resultado.ventas.append(Venta(
                    producto=producto, costo=costo, precio=precio,
                    metodo_pago=metodo, cantidad=cantidad,
                    comision=comision, ganancia_neta=ganancia,
                    notas=notas, fecha=venta_fecha,
                    pagos_combinados=pagos_combinados,
                    hora=hora, vendedor=vendedor,
                    cliente_nombre=cliente_nombre, cliente_cedula=cliente_cedula,
                    cliente_tel=cliente_tel, numero_factura=numero_factura,
                    descuento=descuento, sku=sku,
                    precio_ofertado=precio_ofertado, grupo_venta_id=grupo_venta_id,
                    talla=talla_excel,
                ))
                resultado.meses_afectados.add((venta_fecha.year, venta_fecha.month))
            except ValueError as exc:
                resultado.errores.append(f"Fila {row_idx}: {exc} — omitida")

    # ── Hoja Préstamos ─────────────────────────────────────────────────────
    ws_p = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("préstamos", "prestamos")),
        None,
    )
    if ws_p:
        resultado.prestamos = _leer_prestamos(ws_p)

    # ── Hoja Inventario ────────────────────────────────────────────────────
    ws_i = next(
        (wb[s] for s in wb.sheetnames if s.lower() == "inventario"),
        None,
    )
    if ws_i:
        resultado.productos = _leer_inventario(ws_i)

    # ── Hoja Facturas ──────────────────────────────────────────────────────
    ws_f = next(
        (wb[s] for s in wb.sheetnames if s.lower() == "facturas"),
        None,
    )
    if ws_f:
        resultado.facturas = _leer_facturas(ws_f)

    # ── Hoja Gastos ────────────────────────────────────────────────────────
    ws_g = next(
        (wb[s] for s in wb.sheetnames if s.lower() == "gastos"),
        None,
    )
    if ws_g:
        resultado.gastos = _leer_gastos(ws_g)
        for g in resultado.gastos:
            resultado.meses_gastos_afectados.add((g.fecha.year, g.fecha.month))

    # ── Hoja Configuración ─────────────────────────────────────────────────
    ws_c = next(
        (wb[s] for s in wb.sheetnames
         if s.lower() in ("configuración", "configuracion")),
        None,
    )
    if ws_c:
        resultado.configuracion = _leer_configuracion(ws_c)

    # ── Hoja Notas y Pendientes ────────────────────────────────────────────
    ws_n = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("notas", "notas y pendientes")),
        None,
    )
    if ws_n is not None:
        resultado.notas = _leer_notas(ws_n)

    # ── Hoja Abonos de Facturas ────────────────────────────────────────────
    ws_ab = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("abonos", "abonos de facturas")),
        None,
    )
    if ws_ab is not None:
        resultado.abonos_raw = _leer_abonos(ws_ab)

    # ── Hoja Usuarios ──────────────────────────────────────────────────────
    ws_u = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("usuarios", "users")),
        None,
    )
    if ws_u is not None:
        resultado.usuarios = _leer_usuarios(ws_u)

    # ── Hoja Presupuesto Mensual ────────────────────────────────────────────
    ws_pr = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("presupuesto", "presupuesto mensual")),
        None,
    )
    if ws_pr is not None:
        resultado.presupuestos = _leer_presupuesto(ws_pr)

    # ── Hoja Cuentas ─────────────────────────────────────────────────────────
    ws_cu = next((wb[s] for s in wb.sheetnames if s.lower() == "cuentas"), None)
    if ws_cu is not None:
        resultado.cuentas = _leer_cuentas(ws_cu)

    # ── Hoja Mov. Cuentas ────────────────────────────────────────────────────
    ws_mc = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("mov. cuentas", "mov cuentas")),
        None,
    )
    if ws_mc is not None:
        resultado.movimientos_cuentas = _leer_movimientos_cuentas(ws_mc)

    # ── Hoja Cierres Cuentas ─────────────────────────────────────────────────
    ws_cc = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("cierres cuentas", "cierres")),
        None,
    )
    if ws_cc is not None:
        resultado.cierres_cuentas = _leer_cierres_cuentas(ws_cc)

    # ── Hoja Fiado ───────────────────────────────────────────────────────────
    ws_fi = next((wb[s] for s in wb.sheetnames if s.lower() == "fiado"), None)
    if ws_fi is not None:
        resultado.fiados = _leer_fiado(ws_fi)

    # ── Hoja Abonos Fiado ────────────────────────────────────────────────────
    ws_af = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("abonos fiado", "abonos de apartados")),
        None,
    )
    if ws_af is not None:
        resultado.abonos_fiado_raw = _leer_abonos_fiado(ws_af)

    # ── Hoja Mov. Inventario ─────────────────────────────────────────────────
    ws_mi = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("mov. inventario", "mov inventario")),
        None,
    )
    if ws_mi is not None:
        resultado.movimientos_inventario = _leer_movimientos_inventario(ws_mi)

    # ── Hoja Facturas Items ──────────────────────────────────────────────────
    ws_fit = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("facturas items", "productos por factura")),
        None,
    )
    if ws_fit is not None:
        resultado.facturas_items_raw = _leer_facturas_items(ws_fit)

    # ── Hoja Log Auditoría ───────────────────────────────────────────────────
    ws_log = next(
        (wb[s] for s in wb.sheetnames if s.lower() in ("log auditoría", "log auditoria")),
        None,
    )
    if ws_log is not None:
        resultado.log_acciones = _leer_log_acciones(ws_log)

    return resultado
