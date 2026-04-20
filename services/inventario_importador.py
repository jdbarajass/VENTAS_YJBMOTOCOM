"""
services/inventario_importador.py
Importa productos de inventario desde un archivo .xlsx.

Detecta los encabezados por palabras clave (case-insensitive) para ser
compatible con distintos formatos de Excel del usuario.
"""

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.producto import Producto


def generar_plantilla_inventario(ruta: Path) -> None:
    """
    Genera un archivo .xlsx vacío con los encabezados exactos que espera
    el importador, listo para ser llenado por el usuario.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    lado = Side(style="thin", color="CCCCCC")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)

    # ── Título ────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    titulo = ws["A1"]
    titulo.value = "YJBMOTOCOM — Plantilla de Inventario"
    titulo.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor="1E3A5F")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Subtítulo instructivo ─────────────────────────────────────────
    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = (
        "Llena los datos de cada producto y guarda el archivo. "
        "Luego impórtalo desde el panel Inventario → Importar Excel."
    )
    sub.font = Font(name="Calibri", italic=True, size=10, color="374151")
    sub.fill = PatternFill("solid", fgColor="EFF6FF")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # ── Encabezados ───────────────────────────────────────────────────
    headers = [
        "Numero serial",
        "Producto",
        "Costo unitario",
        "Cantidad actual en bodega: Principal",
        "Codigo de barras",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[3].height = 22

    # ── Filas de ejemplo (gris claro, para guiar) ─────────────────────
    ejemplos = [
        ("001", "Casco X-Sport Rojo T.M",  85000, 5, "7709001234567"),
        ("002", "Aceite 10W-40 1 litro",   18000, 12, ""),
        ("003", "Guantes cuero talla L",    25000, 8, "7709009876543"),
    ]
    for i, (serial, prod, costo, cant, cod) in enumerate(ejemplos, start=4):
        ws.cell(i, 1).value = serial
        ws.cell(i, 2).value = prod
        ws.cell(i, 3).value = costo
        ws.cell(i, 4).value = cant
        ws.cell(i, 5).value = cod
        for col_idx in range(1, 6):
            c = ws.cell(i, col_idx)
            c.fill = PatternFill("solid", fgColor="F1F5F9")
            c.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
            c.border = borde
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 18

    # Nota al pie de los ejemplos
    ws.merge_cells("A7:E7")
    nota = ws["A7"]
    nota.value = "↑ Borra las filas de ejemplo y agrega tus productos desde la fila 4."
    nota.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    nota.fill = PatternFill("solid", fgColor="FFFBEB")
    nota.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 18

    # ── Anchos de columna ─────────────────────────────────────────────
    anchos = [14, 40, 18, 36, 20]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(str(ruta))


# Palabras clave para identificar cada columna
_KW_SERIAL    = {"serial", "numero", "número", "n°", "no.", "#", "consecutivo"}
_KW_PRODUCTO  = {"nombre", "producto", "artículo", "articulo", "descripcion", "descripción", "item"}
_KW_COSTO     = {"costo", "precio", "valor"}
_KW_CANTIDAD  = {"cantidad", "stock", "bodega", "existencia", "unidades"}
_KW_BARRAS    = {"codigo", "código", "barras", "ean", "upc", "barra"}


@dataclass
class ResultadoInventario:
    productos: list[Producto] = field(default_factory=list)
    errores: list[str] = field(default_factory=list)
    total_leidos: int = 0


def _detectar_columnas(fila_headers: list) -> dict[str, int | None]:
    """
    Recibe una lista de valores de la fila de encabezados y retorna
    un dict {campo: índice_columna_0based | None}.
    """
    mapa: dict[str, int | None] = {
        "serial": None, "producto": None, "costo": None,
        "cantidad": None, "codigo_barras": None,
    }

    for idx, cell in enumerate(fila_headers):
        valor = str(cell or "").lower().strip()
        if not valor:
            continue

        if any(kw in valor for kw in _KW_SERIAL):
            if mapa["serial"] is None:
                mapa["serial"] = idx
        if any(kw in valor for kw in _KW_PRODUCTO):
            if mapa["producto"] is None:
                mapa["producto"] = idx
        if any(kw in valor for kw in _KW_COSTO):
            if mapa["costo"] is None:
                mapa["costo"] = idx
        if any(kw in valor for kw in _KW_CANTIDAD):
            if mapa["cantidad"] is None:
                mapa["cantidad"] = idx
        if any(kw in valor for kw in _KW_BARRAS):
            if mapa["codigo_barras"] is None:
                mapa["codigo_barras"] = idx

    return mapa


def importar_inventario_excel(ruta: Path) -> ResultadoInventario:
    """
    Lee un .xlsx con inventario. Detecta automáticamente la fila de encabezados
    buscando la primera fila que contenga palabras clave de columnas.
    """
    resultado = ResultadoInventario()

    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
    except Exception as exc:
        resultado.errores.append(f"No se pudo abrir el archivo: {exc}")
        return resultado

    ws = wb.active
    if ws.max_row < 2:
        resultado.errores.append("El archivo parece estar vacío.")
        return resultado

    # ── Detectar fila de encabezados (buscar en las primeras 5 filas) ──
    header_row_idx = None
    mapa: dict[str, int | None] = {}

    for row_idx in range(1, min(8, ws.max_row + 1)):
        fila = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        # Saltar filas con menos de 2 celdas con contenido (títulos o instrucciones combinadas)
        celdas_con_valor = sum(1 for x in fila if x is not None and str(x).strip())
        if celdas_con_valor < 2:
            continue
        mapa = _detectar_columnas(fila)
        if mapa["producto"] is not None:
            header_row_idx = row_idx
            break

    if header_row_idx is None or mapa["producto"] is None:
        # Sin encabezados detectados — usar posición fija:
        # Col 1=serial, 2=producto, 3=costo, 4=cantidad, 5=barras
        mapa = {"serial": 0, "producto": 1, "costo": 2, "cantidad": 3, "codigo_barras": 4}
        header_row_idx = 1
        resultado.errores.append(
            "No se detectaron encabezados — se usaron las columnas por posición: "
            "1=Serial, 2=Producto, 3=Costo, 4=Cantidad, 5=Código de barras."
        )

    # ── Leer filas de datos ────────────────────────────────────────────
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        # Saltar filas de totales exportadas por el sistema
        if str(ws.cell(row_idx, 1).value or "").strip().upper() == "TOTALES":
            continue

        col_prod = mapa["producto"]
        val_prod = ws.cell(row_idx, col_prod + 1).value if col_prod is not None else None
        if val_prod is None or str(val_prod).strip() == "":
            continue

        resultado.total_leidos += 1
        producto_nombre = str(val_prod).strip()

        def _get(campo: str, default):
            idx = mapa.get(campo)
            if idx is None:
                return default
            return ws.cell(row_idx, idx + 1).value

        # Serial
        serial = str(_get("serial", "") or "").strip()

        # Costo
        try:
            costo = float(_get("costo", 0) or 0)
            if costo < 0:
                costo = 0.0
        except (ValueError, TypeError):
            costo = 0.0

        # Cantidad
        try:
            cantidad = int(float(str(_get("cantidad", 0) or 0).replace(",", ".")))
            if cantidad < 0:
                cantidad = 0
        except (ValueError, TypeError):
            cantidad = 0

        # Código de barras
        cod = str(_get("codigo_barras", "") or "").strip()

        try:
            p = Producto(
                serial=serial,
                producto=producto_nombre,
                costo_unitario=costo,
                cantidad=cantidad,
                codigo_barras=cod,
            )
            resultado.productos.append(p)
        except ValueError as exc:
            resultado.errores.append(f"Fila {row_idx}: {exc} — omitida")

    return resultado
